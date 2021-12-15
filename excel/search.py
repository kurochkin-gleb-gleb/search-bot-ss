import datetime
import time

import openpyxl

from bot.exceptions import NotCorrectColumnType, NotExistColumn
from excel import himera


def parse_row(row, columns_name):
    person = dict()
    for cell, col in zip(row, columns_name):
        person[col] = cell.value
    return person


def parse_excel(sheet):
    rows = sheet.rows
    columns_name = [col.value for col in next(rows)]
    people = []
    for row in rows:
        people.append(parse_row(row, columns_name))
    return people


def get_number(responses):
    phone_numbers = {}
    while len(phone_numbers.values()) != len(responses):
        time.sleep(3)
        for query_id, var in responses:
            res_view = himera.view(query_id)
            if 'error' in res_view:
                phone_numbers[var] = None
            if 'status' in res_view and res_view['status'] == 'ok':
                number = None
                for r in res_view['query']:
                    if 'ТЕЛЕФОН' in r:
                        number = r['ТЕЛЕФОН']
                        break
                    if 'Телефон 1' in r:
                        number = r['Телефон 1']
                        break
                phone_numbers[var] = number
            print(query_id)
        print()
    return phone_numbers


def load_excel(file_name):
    excel = openpyxl.load_workbook(file_name)
    sheet = excel.active
    return excel, sheet, parse_excel(sheet)


def save_excel(excel, file_name):
    *directory, new_file_name = file_name.split('/')
    new_file_name = '/'.join(directory + ['new' + new_file_name])
    excel.save(new_file_name)
    return new_file_name


def search_by_inn(file_name):
    excel, sheet, people = load_excel(file_name)

    responses = []
    for person in people:
        if person['ИНН'] is None:
            continue
        response = himera.search_by_inn(person['ИНН'])
        responses.append((response['query_id'], person['ИНН']))

    inn_to_number = get_number(responses)

    sheet.cell(row=1, column=sheet.max_column + 1).value = 'Номер телефона'
    rows = sheet.rows
    columns_name = [col.value for col in next(rows)]
    for row in rows:
        person = parse_row(row, columns_name)
        row[-1].value = inn_to_number.get(person['ИНН'])

    return save_excel(excel, file_name)


def search_by_passport(file_name):
    excel, sheet, people = load_excel(file_name)

    responses = []
    for person in people:
        if person['Номер документа, удостоверяющего личность'] is None:
            continue
        response = himera.search_by_passport(person['Номер документа, удостоверяющего личность'])
        responses.append((response['query_id'], person['Номер документа, удостоверяющего личность']))

    passport_to_number = get_number(responses)

    sheet.cell(row=1, column=sheet.max_column + 1).value = 'Номер телефона'
    rows = sheet.rows
    columns_name = [col.value for col in next(rows)]
    for row in rows:
        person = parse_row(row, columns_name)
        row[-1].value = passport_to_number.get(person['Номер документа, удостоверяющего личность'])

    return save_excel(excel, file_name)


def search_by_snils(file_name):
    excel, sheet, people = load_excel(file_name)
    if 'снилс' not in people[0]:
        raise NotExistColumn('Колонка «снилс» отсутсвует в таблице')

    responses = []
    for person in people:
        if person['снилс'] is None:
            continue
        response = himera.search_by_snils(person['снилс'])
        responses.append((response['query_id'], person['снилс']))

    snils_to_number = get_number(responses)

    sheet.cell(row=1, column=sheet.max_column + 1).value = 'Номер телефона'
    rows = sheet.rows
    columns_name = [col.value for col in next(rows)]
    for row in rows:
        person = parse_row(row, columns_name)
        row[-1].value = snils_to_number.get(person['снилс'])

    return save_excel(excel, file_name)


def search_by_name(file_name):
    excel, sheet, people = load_excel(file_name)
    if type(people[0]['Дата рождения']) is not datetime.datetime:
        raise NotCorrectColumnType('У даты рождения неверный тип, должна быть дата')

    responses = []
    for person in people:
        response = himera.search_by_name(
            lastname=person['Фамилия'],
            firstname=person['Имя'],
            middlename=person['Отчество'],
            day=person['Дата рождения'].day,
            month=person['Дата рождения'].month,
            year=person['Дата рождения'].year,
        )

        full_name = f"{person['Фамилия']}{person['Имя']}{person['Отчество']}{person['Дата рождения'].day}" \
                    f"{person['Дата рождения'].month}{person['Дата рождения'].year}"
        responses.append((response['query_id'], full_name))

    name_to_number = get_number(responses)

    sheet.cell(row=1, column=sheet.max_column + 1).value = 'Номер телефона'
    rows = sheet.rows
    columns_name = [col.value for col in next(rows)]
    for row in rows:
        person = parse_row(row, columns_name)
        full_name = f"{person['Фамилия']}{person['Имя']}{person['Отчество']}{person['Дата рождения'].day}" \
                    f"{person['Дата рождения'].month}{person['Дата рождения'].year}"
        row[-1].value = name_to_number.get(full_name)

    return save_excel(excel, file_name)
