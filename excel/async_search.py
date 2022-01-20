import asyncio
import datetime

import aiohttp
import openpyxl

from bot.exceptions import NotCorrectColumnType
from excel import async_himera
from bot.handlers import utils


def parse_row(row, columns_name):
    person = dict()
    for cell, col in zip(row, columns_name):
        person[col] = cell.value
    return person


def parse_excel(sheet):
    rows = sheet.rows
    columns_name = [col.value for col in next(rows)]
    people = []
    for row in rows:
        people.append(parse_row(row, columns_name))
    return people


def load_excel(file_name):
    excel = openpyxl.load_workbook(file_name)
    # excel = openpyxl.load_workbook(utils.get_path_to_excel_docs(file_name))
    sheet = excel.active
    return excel, sheet, parse_excel(sheet)


def save_excel(excel, file_name):
    print('*'*40)
    print(file_name)
    new_file_name = utils.get_new_file_name(file_name.split('/')[-1])
    excel.save(utils.file_name_to_file_path(new_file_name))
    # excel.save(new_file_name)
    print(utils.file_name_to_file_path(new_file_name))
    return utils.file_name_to_file_path(new_file_name)


async def get_number(responses, session):
    phone_numbers = {}
    while len(phone_numbers) != len(responses):
        await asyncio.sleep(5)
        tasks = []
        for query_id, var in responses:
            tasks.append(asyncio.ensure_future(async_himera.view(query_id, session, var)))
        for res_view, var in await asyncio.gather(*tasks):
            if var in phone_numbers:
                continue
            if 'error' in res_view:
                phone_numbers[var] = {None: None}
            if 'status' in res_view and res_view['status'] == 'ok':
                number = {}
                if res_view['query'] is None:
                    phone_numbers[var] = {None: None}
                else:
                    for base in res_view['query']:
                        if 'ТЕЛЕФОН' in base and var.startswith(base['ИМЯ'].replace(' ', '')):
                            number[base['БАЗА']] = base['ТЕЛЕФОН']
                    phone_numbers[var] = number
        print('Статистика:', len(phone_numbers), len(responses))
        yield 'statistics', len(phone_numbers), len(responses)
    phone_numbers[None] = {}
    yield 'result', phone_numbers


async def search_by_name(file_name):
    excel, sheet, people = load_excel(file_name)

    if type(people[0]['Дата рождения']) is not datetime.datetime:
        raise NotCorrectColumnType('У даты рождения неверный тип, должна быть дата')

    async with aiohttp.ClientSession() as session:
        tasks = []
        for person in people:
            if person.get('Дата рождения') is None:
                continue
            full_name = f"{person.get('Фамилия')}{person.get('Имя')}{person.get('Отчество')}" \
                        f"{person.get('Дата рождения').day}{person.get('Дата рождения').month}" \
                        f"{person.get('Дата рождения').year}"

            tasks.append(asyncio.ensure_future(async_himera.search_by_name(
                lastname=person.get('Фамилия'),
                firstname=person.get('Имя'),
                middlename=person.get('Отчество'),
                day=person.get('Дата рождения').day,
                month=person.get('Дата рождения').month,
                year=person.get('Дата рождения').year,
                session=session,
                full_name=full_name,
            )))
        responses = []
        for response, full_name in await asyncio.gather(*tasks):
            if 'query_id' not in response:
                print(response)
                err = response[str(response)]
            responses.append((response['query_id'], full_name))
        async for statistics in get_number(responses, session):
            if statistics[0] == 'statistics':
                yield statistics
            else:
                name_to_number = statistics[1]
                break

    number_columns_name = []
    for name, number in name_to_number.items():
        if None in number:
            continue
        for n in number:
            if n in number_columns_name:
                continue
            number_columns_name.append(n)
    number_columns_name.sort(reverse=True, key=lambda x: int(next(filter(lambda y: len(y) == 4 and y.isdigit(), x.split() + ['2000']))))

    used = dict()
    needed_columns = set()
    rows = sheet.rows
    columns_name = [col.value for col in next(rows)]
    for row in rows:
        person = parse_row(row, columns_name)
        if person.get('Дата рождения') is None:
            continue
        full_name = f"{person.get('Фамилия')}{person.get('Имя')}{person.get('Отчество')}" \
                    f"{person.get('Дата рождения').day}{person.get('Дата рождения').month}" \
                    f"{person.get('Дата рождения').year}"

        for name in number_columns_name:
            this_number = name_to_number.get(full_name).get(name)
            if this_number in used.get(full_name, {}).values():
                continue
            if full_name not in used:
                used[full_name] = dict()
            used[full_name][name] = this_number
            needed_columns.add(name)
    number_columns_name = [
        number_column_name for number_column_name in number_columns_name if number_column_name in needed_columns
    ]

    name_to_number_list = dict()
    for name, number in name_to_number.items():
        name_to_number_list[name] = [used.get(name).get(number_column_name)
                                     for number_column_name in number_columns_name if used.get(name) is not None]
        name_to_number_list[name] = list(filter(lambda x: x is not None and len(x) == 10 and x[0] == '9',
                                                name_to_number_list[name]))

    max_column = sheet.max_column
    # for i in range(max(map(len, name_to_number_list.values()))):
    for i in range(3):
        sheet.cell(row=1, column=sheet.max_column + 1).value = f'Телефон {i + 1}'
    rows = sheet.rows
    columns_name = [col.value for col in next(rows)]
    for row in rows:
        person = parse_row(row, columns_name)
        if person.get('Дата рождения') is None:
            continue
        full_name = f"{person.get('Фамилия')}{person.get('Имя')}{person.get('Отчество')}" \
                    f"{person.get('Дата рождения').day}{person.get('Дата рождения').month}" \
                    f"{person.get('Дата рождения').year}"

        for i, number in enumerate(name_to_number_list[full_name]):
            if i == 3:
                break
            row[max_column + i].value = number

    print('FINAL')
    yield 'result', save_excel(excel, file_name)
